from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
import os
import openpyxl
from openpyxl import Workbook

options = webdriver.ChromeOptions()
options.add_argument("disable-extensions")
options.add_argument("disable-plugins")
options.experimental_options["useAutomationExtension"] = False  # prevent load error - Error Loading Extension - Failed to load extension from ... - Could not load extension from ... Loading of unpacked extensions is disabled
driver = webdriver.Chrome(ChromeDriverManager().install(), options=options)

# divides url into 3 parts to loop through the pages
url1 = 'http://www.ageofthering.com/atthemovies/scripts/fellowshipofthering'
url2 = 'to'
url3 = '.php'
# main loop: browses the various pages of the script
# f: first number in page url
rpaste = 1  # paste data in excel
for f in range(1, 38, 4):
    # s: second number in page url
    s = f + 3
    # combines the url
    url = url1 + str(f) + url2 + str(s) + url3
    driver.get(url)
    # finds length of row and column tags on webpage
    rows = len(driver.find_elements_by_xpath("//*[@id='AutoNumber1']/tbody/tr"))
    columns = len(driver.find_elements_by_xpath("//*[@id='AutoNumber1']/tbody/tr[3]/td"))
    # divides url into 3 parts to loop through the rows and columns
    first = "//*[@id='AutoNumber1']/tbody/tr["
    second = "]/td["
    third = "]"
    fname = 'script.xlsx'
    if os.path.exists(fname):
        workbook = openpyxl.load_workbook(fname)
        worksheet = workbook['Sheet']
    else:
        workbook = Workbook()
        worksheet = workbook.active
    # loops through the rows (r) and columns (c) of each page
    # try/except are used to consider rows with only 1 column
    print('Paste Row', rpaste)
    for r in range(1, rows+1):
        while worksheet.cell(rpaste, 1).value:   # get next empty row in sheet
            rpaste += 1
        for c in range(1, columns+1):
            try:
                # combines the xpath of each cell of the table
                final = first + str(r) + second + str(c) + third
                # stores the content of each cell in a variable (data)
                data = driver.find_element_by_xpath(final).text
                if c == 1 and "Scene" in data and "~" in data: # add extra empty row if new scene
                    rpaste += 1
                # writes content of table in an Excel spreadsheet
                worksheet.cell(rpaste, column=c).value = data
            except:
                continue
    workbook.save(fname)
# closes Chrome
driver.quit()
